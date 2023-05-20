import pandas as pd
import re
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, PatternFill

# Create a connection to the SQLite database
engine = create_engine('mysql+pymysql://username:password@localhost:3306/databaseName') # Replace username, password, and databaseName with your info

# Define the column names
columns = ['Name', 'Email', 'Corrected or New Email', 'Remove From Your Organization']

# Define a font for the headers
header_font = Font(bold=True, size=14)

# Define a fill color for the headers (light blue)
header_fill = PatternFill(fill_type="solid", fgColor="87CEFA")

# Query to get the organization ids and names
df_org = pd.read_sql('SELECT id, name FROM organizations', engine) 

# Query to get the organization and create an Excel file for each one
for index, row in df_org.iterrows():
    org_id = row['id']
    org_name = row['name']

    # Sanitize the org_name
    org_name_sanitized = re.sub('[^\w\-_\. ]', '_', org_name)
    org_name_sanitized = org_name_sanitized.replace(' ', '_')


    # Query to get the members of the organization
    query = f'''
    SELECT CONCAT(u.last_name, ', ', u.first_name) AS Name, u.email AS Email
    FROM organization_members om
    JOIN users u ON om.user_id = u.id
    WHERE om.organization_id = {org_id}
    '''
    df = pd.read_sql(query, engine)

    # Create an empty DataFrame with the template columns
    df_template = pd.DataFrame(columns=columns)

    # Fill the 'Name' and 'Email' columns with data from the query
    df_template['Name'] = df['Name']
    df_template['Email'] = df['Email']

    # Write the DataFrame to an Excel file
    file_name = f'{org_name_sanitized}.xlsx'
    df_template.to_excel(file_name, index=False)

    # Load the workbook and select the sheet
    wb = load_workbook(file_name)
    sheet = wb.active

    # Apply the styles to the headers
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Adjust the width of the columns
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook
    wb.save(file_name)